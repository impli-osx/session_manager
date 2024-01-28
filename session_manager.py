import json
import sys
import os
import pandas as pd
import zipfile
import openpyxl
from functools import partial
from PyQt6.QtWidgets import QApplication, QDialog, QLabel, QPushButton, QVBoxLayout, QSpacerItem, QSizePolicy, QLineEdit, QComboBox, QMessageBox
from PyQt6.QtGui import QFont
from PyQt6.QtCore import QTimer, Qt
from ficheentree import FicheWindow as FicheEntreeWindow
from PyQt6.QtWidgets import QApplication
from openpyxl import Workbook
from datetime import datetime
from zipfile import BadZipFile
from openpyxl import load_workbook


# Créer une application Qt
app = QApplication(sys.argv)
# Ne pas quitter l'application lorsque la dernière fenêtre est fermée
app.setQuitOnLastWindowClosed(False)

def load_config():
    try:
        with open('config.json') as f:
            config = json.load(f)
    except FileNotFoundError:
        show_error("Le fichier 'config.json' n'a pas été trouvé dans le répertoire courant.")
        return
    return config



def show_error(message):
    error_dialog = QMessageBox()
    error_dialog.setWindowTitle("Erreur")
    error_dialog.setText(message)
    error_dialog.setIcon(QMessageBox.Icon.Critical)
    error_dialog.exec()
    os._exit(1) 
    
    
global config
config = load_config()
    
    
class FicheEntreeWindow(FicheEntreeWindow):
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.data = {}  # Initialisez self.data dans le constructeur
        self.fields = []
        
        
        if not os.path.isfile('configuration.exe'):
            show_error("Le fichier 'configuration.exe' n'a pas été trouvé dans le répertoire courant.")

    def show_error(self, message):
        error_dialog = QMessageBox()
        error_dialog.setWindowTitle("Erreur")
        error_dialog.setText(message)
        error_dialog.setIcon(QMessageBox.Icon.Critical)
        error_dialog.exec()
        
    # Boucle les champs présents dans FicheEntreeWindow
    def update_data(self):
        fields = self.findChildren(QLineEdit)
        for field in fields:
            self.data[field.objectName()] = field.text()
        combos = self.findChildren(QComboBox)
        for combo in combos:
            if combo.objectName() == "session_duration_combo":
                self.data["ChoixDuréeSession"] = combo.currentText()
    
    
    
    # Récupère les valeurs des différents champs
    def get_data(self):
        return self.data
    
    
    # Ajuste la largeur des colonnes du fichier Excel
    def adjust_column_widths(self, sheet):
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width



    # Enregistre les données dans le fichier Excel
    def save_to_excel(self, data):
        current_year = str(datetime.now().year)
        with open('fiche.json', 'r') as f:
            champs = json.load(f)
        ordered_keys = [champ.get("label_content", "").replace(" ", "") for order, champ in sorted(champs.items(), key=lambda item: int(item[0]))]
        ordered_keys.append("ChoixDuréeSession")
        new_data_df = pd.DataFrame(data, columns=ordered_keys, index=[0])
        if not os.path.isfile('data.xlsx'):
            new_data_df.to_excel('data.xlsx', sheet_name=current_year, index=False)
        else:
            book = load_workbook('data.xlsx')
            if current_year in book.sheetnames:
                data_df = pd.read_excel('data.xlsx', sheet_name=current_year, engine='openpyxl')
                data_df = pd.concat([data_df, new_data_df], ignore_index=True)
            else:
                data_df = new_data_df
            # Create a new Excel file with the updated data
            with pd.ExcelWriter('data_temp.xlsx', engine='openpyxl') as writer:
                data_df.to_excel(writer, index=False, sheet_name=current_year)
            # Replace the old file with the new one
            os.remove('data.xlsx')
            os.rename('data_temp.xlsx', 'data.xlsx')
        book = load_workbook("data.xlsx")
        sheet = book[current_year]
        self.adjust_column_widths(sheet)
        book.save("data.xlsx")
    
    
    
    # On déclenche les timers à la fermeture de la fenêtre FicheEntreeWindow
    def closeEvent(self, event):
        super().closeEvent(event)
        timer_duree_session()
        timer_popup_2()
        if config['fermeture']['fermeture_popup']:
            timer_fermeture()
        # Si le log est activé, enregistrer les données dans le fichier Excel
        if config['fiche']['fiche_log']:
            self.update_data()
            #print(self.data) # Pour débugger
            self.save_to_excel(self.get_data())
        
        


# Fonction pour créer le popup d'après le fichier de configuration
def creation_popup(titre, texte, largeur, hauteur, police, taille_police, delai, fullscreen=False, use_timer=True):
    #print(f"Appel fonction popup : {titre}, {texte}, {largeur}, {hauteur}, {police}, {taille_police}, {delai}") # Pour débugger
    global fenetre
    fenetre = QDialog()
    fenetre.resize(largeur, hauteur)
    fenetre.setWindowTitle(titre)
    fenetre.setStyleSheet("background-color: white; border-radius: 10px;")
    layout = QVBoxLayout(fenetre)
    if fullscreen:
        # Ajouter un QSpacerItem en haut pour centrer le texte
        layout.addItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))
    label = QLabel(texte)
    font = QFont(police, taille_police)
    label.setFont(font)
    if fullscreen:
        fenetre.setStyleSheet("background-color: white;")
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
    layout.addWidget(label)
    if fullscreen:
        # Ajouter un QSpacerItem en bas pour centrer le texte
        layout.addItem(QSpacerItem(20, 40, QSizePolicy.Policy.Minimum, QSizePolicy.Policy.Expanding))
    if use_timer:
        bouton = QPushButton("J'ai compris")
        bouton.setFont(font)  # Utiliser la même police que le label
        bouton.setStyleSheet(f"font-size: {taille_police}px;")  # Définir la taille de la police
        bouton.clicked.connect(fenetre.close)
        layout.addWidget(bouton)
    fenetre.setLayout(layout)
    if use_timer:
        QTimer.singleShot(delai * 1000, fenetre.close)
    if fullscreen:
        fenetre.showFullScreen()
    else:
        fenetre.show()
    #print("Popup créée.") # Pour débugger



# Fonction pour créer le timer de la durée de la session
def timer_duree_session():
    # Convertir la durée de la session en secondes
    duree_session = (int(config['timer']['duree_session']) * 60) + int(config['timer']['timer_popup_3'])
    #print(f"Durée session : {duree_session}") # Pour débugger
    # Créer un timer
    global timer
    timer = QTimer()
    # Connecter le signal timeout du timer à la fonction de fermeture de session
    timer.timeout.connect(end_session)
    # Démarrer le timer
    timer.start(duree_session * 1000)  # Convertir les secondes en millisecondes
    # Imprimer un message pour indiquer que le timer est actif
    #print("Le timer durée session est actif.") # Pour débugger
    timer_popup_1()



# Fonction pour créer le timer du premier popup
def timer_popup_1():
    timer_1 = (int(config['timer']['timer_popup_1']) * 60)
    global timer1
    #print(f"Timer popup 1 : {timer_1}") # Pour débugger
    timer1 = QTimer(app)
    timer1.stop()  # Arrêter le timer avant de le configurer
    creation_popup_1 = partial(creation_popup, config['titre']['titre_popup_1'], config['text']['text_popup_1'], int(config['style']['largeur_popup']), int(config['style']['hauteur_popup']), config['style']['police'], int(config['style']['taille_police']), int(config['timer']['delai_fermeture']))
    timer1.timeout.connect(creation_popup_1)
    timer1.timeout.connect(timer1.stop)
    timer1.start(timer_1 * 1000) # DEBUG
    #print("Le timer popup 1 est actif.") # Pour débugger



# Fonction pour créer le timer du deuxième popup
def timer_popup_2():
    global timer_2
    timer_2 = (int(config['timer']['duree_session']) * 60) - (int(config['timer']['timer_popup_2']) * 60)
    #print(f"Timer popup 2 : {timer_2}") # Pour débugger
    global timer2
    timer2 = QTimer(app)
    timer2.stop()  # Arrêter le timer avant de le configurer
    creation_popup_2 = partial(creation_popup, config['titre']['titre_popup_2'], config['text']['text_popup_2'], int(config['style']['largeur_popup']), int(config['style']['hauteur_popup']), config['style']['police'], int(config['style']['taille_police']), int(config['timer']['delai_fermeture']))
    timer2.timeout.connect(creation_popup_2)
    timer2.timeout.connect(timer2.stop)
    timer2.start(timer_2 * 1000) # Convertir les secondes en millisecondes
    #print("Le timer popup 2 est actif.") # Pour débugger
    
    

# Fonction pour créer le timer de fermeture de la session
def timer_fermeture():
    timerfermeture = (int(config['timer']['duree_session']) * 60)
    #print(f"Timer fermeture : {timerfermeture}") # Pour débugger
    global timerfin
    timerfin = QTimer(app)
    timerfin.stop()  # Arrêter le timer avant de le configurer
    creation_fin = partial(creation_popup, config['titre']['titre_popup_3'], config['text']['text_popup_3'], int(config['style']['largeur_popup']), int(config['style']['hauteur_popup']), config['style']['police'], int(config['style']['taille_police']), int(config['timer']['delai_fermeture']), fullscreen=True, use_timer=False)
    timerfin.timeout.connect(creation_fin)
    timerfin.timeout.connect(timerfin.stop)
    timerfin.start(timerfermeture * 1000)
    
    

# Fonction appelée par duree_session : déconnecte ou verrouille la session
def end_session():
    # Fermer la session en fonction de la valeur de session_fermeture
    if config['fermeture']['fermeture_session'] == 'Déconnecter':
        # Déconnecter l'utilisateur
        os.system('logoff')
        #print("Fin de session. Déconnexion.") # Pour débugger
    else:
        # Verrouiller la session
        os.system('rundll32.exe user32.dll,LockWorkStation')
        #print("Fin de session. Verrouillage.") # Pour débugger
    app.quit() # Terminer l'application Qt



# # Créer une application Qt
# app = QApplication(sys.argv)
# # Ne pas quitter l'application lorsque la dernière fenêtre est fermée
# app.setQuitOnLastWindowClosed(False)
# Créer et afficher la fenêtre FicheEntreeWindow
window = FicheEntreeWindow()
window.showFullScreen()
# Démarrer la boucle d'événements
sys.exit(app.exec())